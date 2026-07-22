import io
import re

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from llm.schema import ExtractedDocument

INVALID_SHEET_NAME_CHARS = re.compile(r"[\\/?*\[\]:]")

SCALAR_FIELDS = [
    "document_type",
    "invoice_number",
    "invoice_date",
    "due_date",
    "vendor_name",
    "vendor_address",
    "customer_name",
    "subtotal",
    "tax",
    "total",
    "currency",
]

LINE_ITEM_HEADERS = ["description", "quantity", "unit_price", "amount"]

MAX_SHEET_NAME_LENGTH = 31


def _unique_sheet_name(workbook: Workbook, title: str) -> str:
    cleaned = INVALID_SHEET_NAME_CHARS.sub(" ", title or "").strip(" '") or "Table"
    base = cleaned[:MAX_SHEET_NAME_LENGTH]
    name = base
    suffix = 2
    while name in workbook.sheetnames:
        suffix_str = f" ({suffix})"
        name = base[: MAX_SHEET_NAME_LENGTH - len(suffix_str)] + suffix_str
        suffix += 1
    return name


def build_workbook(data: ExtractedDocument) -> bytes:
    workbook = Workbook()

    summary_sheet: Worksheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Field", "Value"])

    if data.summary:
        summary_sheet.append(["summary", data.summary])
    for field in SCALAR_FIELDS:
        value = getattr(data, field)
        if value is not None:
            summary_sheet.append([field, value])
    for extra in data.fields:
        summary_sheet.append([extra.label, extra.value])

    if data.line_items:
        line_items_sheet = workbook.create_sheet("Line Items")
        line_items_sheet.append(LINE_ITEM_HEADERS)
        for item in data.line_items:
            line_items_sheet.append([getattr(item, header) for header in LINE_ITEM_HEADERS])

    for table in data.tables:
        sheet = workbook.create_sheet(_unique_sheet_name(workbook, table.title))
        if table.headers:
            sheet.append(table.headers)
        for row in table.rows:
            sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
